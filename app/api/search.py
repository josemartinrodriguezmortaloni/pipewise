# app/api/search.py - Rutas para búsqueda y exportación
from fastapi import APIRouter, HTTPException, status, Depends, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from typing import List, Optional, Dict, Any
import logging
import csv
import json
import io
from datetime import datetime, date

from app.auth.middleware import get_current_user, get_manager_user
from app.models.user import User
from app.schemas.crm_schema import (
    SearchRequest,
    SearchResponse,
    SearchResult,
    ExternalLeadImport,
    ImportResult,
)

logger = logging.getLogger(__name__)

# Crear router para búsqueda
router = APIRouter(prefix="/search", tags=["Search & Export"])


# ===================== BÚSQUEDA AVANZADA =====================


@router.post("/", response_model=SearchResponse)
async def search_entities(
    search_data: SearchRequest,
    current_user: User = Depends(get_current_user),
):
    """Búsqueda avanzada en todas las entidades"""
    try:
        start_time = datetime.utcnow()

        # Implementar búsqueda real aquí
        # 1. Preparar query de búsqueda
        # 2. Buscar en leads, contactos, oportunidades
        # 3. Aplicar filtros de permisos del usuario
        # 4. Rankear resultados por relevancia

        # Simular resultados de búsqueda
        results = []

        # Ejemplo de resultados simulados
        if "john" in search_data.query.lower():
            results.append(
                SearchResult(
                    entity_type="leads",
                    entity_id="lead_123",
                    title="John Smith - Acme Corp",
                    description="Potential customer interested in our services",
                    score=0.95,
                    highlights={
                        "name": ["<em>John</em> Smith"],
                        "company": ["Acme Corp"],
                    },
                )
            )

        if "acme" in search_data.query.lower():
            results.append(
                SearchResult(
                    entity_type="opportunities",
                    entity_id="opp_456",
                    title="Acme Corp - Enterprise Deal",
                    description="$50,000 enterprise software opportunity",
                    score=0.87,
                    highlights={
                        "company": ["<em>Acme</em> Corp"],
                        "amount": ["$50,000"],
                    },
                )
            )

        end_time = datetime.utcnow()
        took = (end_time - start_time).total_seconds()

        return SearchResponse(
            results=results, total=len(results), query=search_data.query, took=took
        )

    except Exception as e:
        logger.error(f"Search error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Search failed"
        )


@router.get("/suggestions")
async def get_search_suggestions(
    q: str = Query(..., min_length=2, description="Query para sugerencias"),
    limit: int = Query(10, ge=1, le=20, description="Límite de sugerencias"),
    current_user: User = Depends(get_current_user),
):
    """Obtener sugerencias de búsqueda"""
    try:
        # Implementar sugerencias de autocompletado
        suggestions = []

        # Simular sugerencias basadas en la query
        if q.lower().startswith("j"):
            suggestions.extend(
                [
                    {"type": "lead", "text": "John Smith", "entity_id": "lead_123"},
                    {
                        "type": "company",
                        "text": "Johnson & Associates",
                        "entity_id": "comp_456",
                    },
                ]
            )

        if q.lower().startswith("a"):
            suggestions.extend(
                [
                    {
                        "type": "company",
                        "text": "Acme Corporation",
                        "entity_id": "comp_789",
                    },
                    {"type": "lead", "text": "Alice Johnson", "entity_id": "lead_234"},
                ]
            )

        return {"suggestions": suggestions[:limit], "query": q}

    except Exception as e:
        logger.error(f"Search suggestions error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get search suggestions",
        )


@router.get("/recent")
async def get_recent_searches(
    limit: int = Query(10, ge=1, le=50),
    current_user: User = Depends(get_current_user),
):
    """Obtener búsquedas recientes del usuario"""
    try:
        # Implementar historial de búsquedas
        recent_searches = [
            {
                "query": "john smith",
                "timestamp": datetime.utcnow().isoformat(),
                "results_count": 3,
            },
            {
                "query": "acme corp",
                "timestamp": datetime.utcnow().isoformat(),
                "results_count": 7,
            },
        ]

        return {"recent_searches": recent_searches[:limit]}

    except Exception as e:
        logger.error(f"Recent searches error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get recent searches",
        )


# ===================== EXPORTACIÓN DE DATOS =====================


@router.get("/export/leads")
async def export_leads(
    format: str = Query("csv", regex="^(csv|json|xlsx)$"),
    status_filter: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(get_current_user),
):
    """Exportar leads en diferentes formatos"""
    try:
        # Aplicar filtros y obtener datos
        leads_data = get_filtered_leads(
            user=current_user,
            status_filter=status_filter,
            date_from=date_from,
            date_to=date_to,
        )

        if format == "csv":
            return export_to_csv(leads_data, "leads")
        elif format == "json":
            return export_to_json(leads_data)
        elif format == "xlsx":
            return export_to_xlsx(leads_data, "leads")

    except Exception as e:
        logger.error(f"Export leads error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export leads",
        )


@router.get("/export/opportunities")
async def export_opportunities(
    format: str = Query("csv", regex="^(csv|json|xlsx)$"),
    stage_filter: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: User = Depends(get_current_user),
):
    """Exportar oportunidades"""
    try:
        # Obtener datos de oportunidades
        opportunities_data = get_filtered_opportunities(
            user=current_user,
            stage_filter=stage_filter,
            date_from=date_from,
            date_to=date_to,
        )

        if format == "csv":
            return export_to_csv(opportunities_data, "opportunities")
        elif format == "json":
            return export_to_json(opportunities_data)
        elif format == "xlsx":
            return export_to_xlsx(opportunities_data, "opportunities")

    except Exception as e:
        logger.error(f"Export opportunities error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export opportunities",
        )


@router.get("/export/contacts")
async def export_contacts(
    format: str = Query("csv", regex="^(csv|json|xlsx)$"),
    current_user: User = Depends(get_current_user),
):
    """Exportar contactos"""
    try:
        # Obtener datos de contactos
        contacts_data = get_filtered_contacts(user=current_user)

        if format == "csv":
            return export_to_csv(contacts_data, "contacts")
        elif format == "json":
            return export_to_json(contacts_data)
        elif format == "xlsx":
            return export_to_xlsx(contacts_data, "contacts")

    except Exception as e:
        logger.error(f"Export contacts error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export contacts",
        )


# ===================== IMPORTACIÓN DE DATOS =====================


@router.post("/import/leads", response_model=ImportResult)
async def import_leads(
    import_data: ExternalLeadImport,
    background_tasks: BackgroundTasks,
    manager_user: User = Depends(get_manager_user),
):
    """Importar leads desde fuentes externas"""
    try:
        # Validar datos de importación
        if not import_data.leads:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="No leads data provided"
            )

        # Procesar importación en background
        import_id = f"import_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        background_tasks.add_task(
            process_leads_import, import_data, import_id, str(manager_user.id)
        )

        return ImportResult(
            total_records=len(import_data.leads),
            successful_imports=0,  # Se actualizará en background
            failed_imports=0,
            errors=[],
            import_id=import_id,
            status="processing",
        )

    except Exception as e:
        logger.error(f"Import leads error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to start import process",
        )


@router.get("/import/status/{import_id}")
async def get_import_status(
    import_id: str,
    current_user: User = Depends(get_current_user),
):
    """Obtener estado de importación"""
    try:
        # Implementar consulta de estado de importación
        # Por ahora simular respuesta
        return {
            "import_id": import_id,
            "status": "completed",
            "total_records": 100,
            "successful_imports": 95,
            "failed_imports": 5,
            "progress": 100,
            "created_at": datetime.utcnow().isoformat(),
            "completed_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.error(f"Get import status error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get import status",
        )


# ===================== FUNCIONES AUXILIARES =====================


def get_filtered_leads(
    user: User,
    status_filter: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> List[dict]:
    """Obtener leads filtrados"""
    # Implementar consulta real a la base de datos
    # Por ahora simular datos
    return [
        {
            "id": "lead_123",
            "name": "John Smith",
            "email": "john@example.com",
            "company": "Acme Corp",
            "status": "qualified",
            "created_at": datetime.utcnow().isoformat(),
        },
        {
            "id": "lead_456",
            "name": "Jane Doe",
            "email": "jane@company.com",
            "company": "Tech Solutions",
            "status": "new",
            "created_at": datetime.utcnow().isoformat(),
        },
    ]


def get_filtered_opportunities(
    user: User,
    stage_filter: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> List[dict]:
    """Obtener oportunidades filtradas"""
    return [
        {
            "id": "opp_123",
            "name": "Enterprise Deal",
            "amount": 50000,
            "stage": "proposal",
            "probability": 75,
            "close_date": date.today().isoformat(),
        }
    ]


def get_filtered_contacts(user: User) -> List[dict]:
    """Obtener contactos filtrados"""
    return [
        {
            "id": "contact_123",
            "first_name": "John",
            "last_name": "Smith",
            "email": "john@example.com",
            "company": "Acme Corp",
            "position": "CEO",
        }
    ]


def export_to_csv(data: List[dict], entity_type: str) -> StreamingResponse:
    """Exportar datos a CSV"""
    output = io.StringIO()

    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    output.seek(0)

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={entity_type}_{datetime.now().strftime('%Y%m%d')}.csv"
        },
    )


def export_to_json(data: List[dict]) -> StreamingResponse:
    """Exportar datos a JSON"""
    json_data = json.dumps(data, indent=2, default=str)

    return StreamingResponse(
        io.BytesIO(json_data.encode()),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=export_{datetime.now().strftime('%Y%m%d')}.json"
        },
    )


def export_to_xlsx(data: List[dict], entity_type: str) -> StreamingResponse:
    """Exportar datos a Excel (requiere openpyxl)"""
    # Nota: Esta función requiere la librería openpyxl
    # pip install openpyxl

    try:
        import openpyxl
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = entity_type.capitalize()

        if data:
            # Escribir headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Escribir datos
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=str(item.get(header, "")))

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={entity_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            },
        )

    except ImportError:
        # Si openpyxl no está disponible, retornar CSV
        logger.warning("openpyxl not available, falling back to CSV export")
        return export_to_csv(data, entity_type)


async def process_leads_import(
    import_data: ExternalLeadImport, import_id: str, user_id: str
):
    """Procesar importación de leads en background"""
    try:
        logger.info(f"Starting import process {import_id} for user {user_id}")

        successful = 0
        failed = 0
        errors = []

        for i, lead_data in enumerate(import_data.leads):
            try:
                # Implementar lógica de importación
                # 1. Validar datos del lead
                # 2. Mapear campos según mapping
                # 3. Crear lead en la base de datos
                # 4. Manejar duplicados

                # Simular procesamiento
                if "email" in lead_data and "@" in lead_data["email"]:
                    successful += 1
                else:
                    failed += 1
                    errors.append(
                        {
                            "row": i + 1,
                            "error": "Invalid email format",
                            "data": lead_data,
                        }
                    )

            except Exception as e:
                failed += 1
                errors.append({"row": i + 1, "error": str(e), "data": lead_data})

        # Actualizar estado de importación en la base de datos
        logger.info(
            f"Import {import_id} completed: {successful} successful, {failed} failed"
        )

    except Exception as e:
        logger.error(f"Import process {import_id} failed: {e}")
